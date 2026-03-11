"""
process_data.py
Run this script each time you receive a new monthly Excel export.
It reads Order_Listing_*.xlsx (or specify a path) and outputs dashboard_data.json.

Usage:
    python3 process_data.py                          # auto-detects newest .xlsx in current dir
    python3 process_data.py path/to/orders.xlsx      # specify file
"""

import sys
import json
import glob
import os
from collections import defaultdict, Counter
from datetime import datetime

# ── helpers ──────────────────────────────────────────────────────────────────

MONTH_ORDER_KEY = lambda x: datetime.strptime(x, "%b %Y")
DOW_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
TIME_SLOTS = [
    "00:00–04:00",
    "04:00–08:00",
    "08:00–12:00",
    "12:00–16:00",
    "16:00–20:00",
    "20:00–24:00",
]
# Items to exclude (packaging/misc charges that appear as menu line items)
EXCLUDE_ITEMS = {"packing", "packing (t)", "container charge", "delivery charge"}


def get_slot(hour):
    if hour < 4:   return "00:00–04:00"
    if hour < 8:   return "04:00–08:00"
    if hour < 12:  return "08:00–12:00"
    if hour < 16:  return "12:00–16:00"
    if hour < 20:  return "16:00–20:00"
    return "20:00–24:00"


def quartiles(vals):
    s = sorted(vals)
    n = len(s)
    if n == 0:
        return dict(
            min=0, q1=0, median=0, q3=0, max=0, mean=0, n=0,
            whisker_lower=0, whisker_upper=0, outliers=[],
        )
    def perc(p):
        idx = (n - 1) * p
        lo, hi = int(idx), min(int(idx) + 1, n - 1)
        return s[lo] + (s[hi] - s[lo]) * (idx - lo)
    q1 = perc(0.25)
    q3 = perc(0.75)
    iqr = q3 - q1
    fence_lo = q1 - 1.5 * iqr
    fence_hi = q3 + 1.5 * iqr
    inliers = [v for v in s if fence_lo <= v <= fence_hi]
    whisker_lower = round(min(inliers) if inliers else q1, 2)
    whisker_upper = round(max(inliers) if inliers else q3, 2)
    outliers = sorted(set(round(v, 2) for v in s if v < fence_lo or v > fence_hi))
    return dict(
        min=round(s[0], 2),
        q1=round(q1, 2),
        median=round(perc(0.50), 2),
        q3=round(q3, 2),
        max=round(s[-1], 2),
        mean=round(sum(s) / n, 2),
        n=n,
        whisker_lower=whisker_lower,
        whisker_upper=whisker_upper,
        outliers=outliers,
    )


# ── load data ─────────────────────────────────────────────────────────────────

DRINK_CATEGORIES = {
    "coffee", "cold coffee and brews", "tea", "mocktails",
    "shakes", "mojito", "smoothie", "fresh for you",
}

def load_item_sales(path):
    """Read Petpooja Item Wise Sales Report → dict of {name: {qty, revenue, category}}."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Run: pip3 install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_idx = next(
        (i for i, r in enumerate(rows) if r and r[0] == 'Category'), None
    )
    if header_idx is None:
        raise ValueError(f"Could not find header row in {path}")

    # Aggregate qty + revenue by item name (same item can appear under multiple codes)
    item_data = defaultdict(lambda: dict(qty=0.0, revenue=0.0, category=''))

    skip = {'Sub Total', 'Total', 'Min.', 'Max.', 'Avg.', 'Category', None, ''}
    current_cat = ''
    for r in rows[header_idx + 1:]:
        if not r or len(r) < 6:
            continue
        category, name, code, sap, qty, revenue = r[0], r[1], r[2], r[3], r[4], r[5]
        if category and category not in skip:
            current_cat = category
        if category == 'Sub Total' or name in (None, ''):
            continue
        if not isinstance(qty, (int, float)) or not isinstance(revenue, (int, float)):
            continue
        if name.lower() in EXCLUDE_ITEMS:
            continue
        item_data[name]['qty'] += qty
        item_data[name]['revenue'] += revenue
        if not item_data[name]['category']:
            item_data[name]['category'] = current_cat

    return item_data

def find_excel():
    """Return a list of Order_Listing Excel paths to load (merged if multiple)."""
    if len(sys.argv) > 1:
        return sys.argv[1:]
    files = sorted(glob.glob("Order_Listing_*.xlsx"), key=os.path.getmtime)
    if not files:
        # Fallback: any xlsx that isn't an item sales report
        files = sorted(
            [f for f in glob.glob("*.xlsx") if not f.startswith("Item_Wise")],
            key=os.path.getmtime,
        )
    if not files:
        raise FileNotFoundError("No Order_Listing_*.xlsx file found in current directory.")
    if len(files) > 1:
        print(f"Found {len(files)} order listing files — merging all:")
        for f in files:
            print(f"  {f}")
    else:
        print(f"Using: {files[0]}")
    return files


def load_orders(path):
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Run: pip3 install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (row with 'Order No.')
    header_idx = next(
        i for i, r in enumerate(rows) if r and r[0] == "Order No."
    )
    headers = rows[header_idx]
    col = {h: i for i, h in enumerate(headers) if h}

    parsed = []
    for r in rows[header_idx + 1 :]:
        if r[col.get("Status", 20)] != "Printed":
            continue
        created_raw = r[col.get("Created", 21)]
        if not created_raw:
            continue
        try:
            dt = datetime.strptime(str(created_raw).strip(), "%d %b %Y %H:%M:%S")
        except ValueError:
            continue
        parsed.append(
            dict(
                dt=dt,
                total=float(r[col.get("Grand Total (₹)", 17)] or 0),
                items_str=str(r[col.get("Items", 10)] or ""),
            )
        )
    return parsed


# ── aggregations ──────────────────────────────────────────────────────────────

def build_monthly_revenue(orders):
    monthly = defaultdict(lambda: dict(revenue=0.0, orders=0))
    for o in orders:
        key = o["dt"].strftime("%b %Y")
        monthly[key]["revenue"] += o["total"]
        monthly[key]["orders"] += 1
    months = sorted(monthly.keys(), key=MONTH_ORDER_KEY)
    return [
        dict(month=m, revenue=round(monthly[m]["revenue"], 2), orders=monthly[m]["orders"])
        for m in months
    ]


def build_dow_distributions(orders):
    # Map: dow -> date_str -> daily_total
    day_daily = {d: defaultdict(float) for d in DOW_ORDER}
    for o in orders:
        dow = o["dt"].strftime("%A")
        date_str = o["dt"].strftime("%Y-%m-%d")
        day_daily[dow][date_str] += o["total"]

    result = {}
    for dow in DOW_ORDER:
        vals = list(day_daily[dow].values())
        result[dow] = dict(
            stats=quartiles(vals),
            values=[round(v, 2) for v in sorted(vals)],
        )
    return result


def build_top_items(orders, top_n=20):
    # Count item appearances across all orders (by month)
    item_monthly = defaultdict(lambda: defaultdict(int))
    item_total = Counter()
    item_order_totals = defaultdict(list)  # order values when this item appeared

    months_seen = set()
    for o in orders:
        month = o["dt"].strftime("%b %Y")
        months_seen.add(month)
        if o["items_str"]:
            items = [i.strip() for i in o["items_str"].split(",") if i.strip()]
            for item in items:
                if item.lower() in EXCLUDE_ITEMS:
                    continue
                item_monthly[item][month] += 1
                item_total[item] += 1
                item_order_totals[item].append(o["total"])

    months = sorted(months_seen, key=MONTH_ORDER_KEY)
    result = []
    for item, total_count in item_total.most_common(top_n):
        monthly_counts = {m: item_monthly[item].get(m, 0) for m in months}
        avg_per_month = round(total_count / len(months), 1)
        totals = item_order_totals[item]
        avg_order_value = round(sum(totals) / len(totals), 2) if totals else 0
        result.append(
            dict(
                name=item,
                total_count=total_count,
                avg_per_month=avg_per_month,
                avg_order_value=avg_order_value,
                monthly=monthly_counts,
            )
        )
    return result, months


def build_monthly_detail(orders, top_n=10):
    """Per-month breakdown: daily revenue/orders + top N items for that month."""
    # Group orders by month
    month_orders = defaultdict(list)
    for o in orders:
        key = o["dt"].strftime("%b %Y")
        month_orders[key].append(o)

    result = {}
    for month, mos in month_orders.items():
        # Daily revenue
        daily = defaultdict(lambda: dict(revenue=0.0, orders=0))
        item_counts = Counter()
        for o in mos:
            day_str = o["dt"].strftime("%Y-%m-%d")
            daily[day_str]["revenue"] += o["total"]
            daily[day_str]["orders"] += 1
            if o["items_str"]:
                for item in [i.strip() for i in o["items_str"].split(",") if i.strip()]:
                    if item.lower() not in EXCLUDE_ITEMS:
                        item_counts[item] += 1

        days_sorted = sorted(daily.keys())
        result[month] = dict(
            daily=[
                dict(
                    date=d,
                    revenue=round(daily[d]["revenue"], 2),
                    orders=daily[d]["orders"],
                )
                for d in days_sorted
            ],
            top_items=[
                dict(name=item, count=cnt)
                for item, cnt in item_counts.most_common(top_n)
            ],
        )
    return result


def _build_ranked_list(item_data, sort_key, top_n, food_only=False):
    filtered = {
        name: d for name, d in item_data.items()
        if not food_only or d['category'].lower() not in DRINK_CATEGORIES
    }
    sorted_items = sorted(filtered.items(), key=lambda x: x[1][sort_key], reverse=True)
    result = []
    for name, d in sorted_items[:top_n]:
        qty, revenue = d['qty'], d['revenue']
        result.append(dict(
            name=name,
            qty=round(qty, 1),
            revenue=round(revenue, 2),
            avg_price=round(revenue / qty, 2) if qty else 0,
            category=d['category'],
        ))
    return result


def build_top_items_by_revenue(item_data, top_n=20):
    return _build_ranked_list(item_data, 'revenue', top_n)

def build_top_items_by_qty(item_data, top_n=20):
    return _build_ranked_list(item_data, 'qty', top_n)

def build_top_food_by_revenue(item_data, top_n=20):
    return _build_ranked_list(item_data, 'revenue', top_n, food_only=True)

def build_top_food_by_qty(item_data, top_n=20):
    return _build_ranked_list(item_data, 'qty', top_n, food_only=True)


def find_item_sales_report():
    files = sorted(glob.glob("Item_Wise_Sales_Report_*.xlsx"), key=os.path.getmtime, reverse=True)
    return files[0] if files else None


def build_time_slots(orders):
    slot_data = {s: dict(revenue=0.0, orders=0) for s in TIME_SLOTS}
    for o in orders:
        slot = get_slot(o["dt"].hour)
        slot_data[slot]["revenue"] += o["total"]
        slot_data[slot]["orders"] += 1
    return [
        dict(
            slot=s,
            revenue=round(slot_data[s]["revenue"], 2),
            orders=slot_data[s]["orders"],
            label=s,
        )
        for s in TIME_SLOTS
    ]


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    paths = find_excel()
    print("Loading orders…")
    orders = []
    seen_keys = set()
    for path in paths:
        batch = load_orders(path)
        # Deduplicate across files by (datetime, total) key
        for o in batch:
            key = (o["dt"], o["total"])
            if key not in seen_keys:
                seen_keys.add(key)
                orders.append(o)
    print(f"  {len(orders)} printed orders loaded.")

    top_items, months = build_top_items(orders)

    item_sales_path = find_item_sales_report()
    if item_sales_path:
        print(f"  Loading item sales: {item_sales_path}")
        item_data = load_item_sales(item_sales_path)
        top_by_revenue = build_top_items_by_revenue(item_data)
        top_by_qty = build_top_items_by_qty(item_data)
        top_food_by_revenue = build_top_food_by_revenue(item_data)
        top_food_by_qty = build_top_food_by_qty(item_data)
        print(f"  Food items: {sum(1 for d in item_data.values() if d['category'].lower() not in DRINK_CATEGORIES)}, Drink items: {sum(1 for d in item_data.values() if d['category'].lower() in DRINK_CATEGORIES)}")
        print(f"  {len(item_data)} unique items loaded from item sales report.")
    else:
        print("  No Item_Wise_Sales_Report_*.xlsx found — skipping item revenue data.")
        top_by_revenue = []
        top_by_qty = []
        top_food_by_revenue = []
        top_food_by_qty = []

    data = dict(
        restaurant="No Strings Attached",
        address="BH 24 Salt Lake City Sector 2",
        period=f"{months[0]} – {months[-1]}",
        generated=datetime.now().strftime("%d %b %Y %H:%M"),
        months=months,
        monthly_revenue=build_monthly_revenue(orders),
        dow_distributions=build_dow_distributions(orders),
        top_items=top_items,
        top_items_by_revenue=top_by_revenue,
        top_items_by_qty=top_by_qty,
        top_food_by_revenue=top_food_by_revenue,
        top_food_by_qty=top_food_by_qty,
        time_slots=build_time_slots(orders),
        monthly_detail=build_monthly_detail(orders),
    )

    out_path = "dashboard_data.json"
    with open(out_path, "w") as f:
        json.dump(data, f, indent=2)
    print(f"  Written → {out_path}")


if __name__ == "__main__":
    main()
